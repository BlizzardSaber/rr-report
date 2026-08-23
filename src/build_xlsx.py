"""按现有《RR API数据.xlsx》的样式生成报表（v3）。

工作表结构：
- 「分配数据」：全部累积的分配明细，带「班次」列，整体按时间倒序（最原始的排序）
- 「agent上下线数据-夜班」：夜班名单内客服的会话，按上线时间倒序
- 「agent上下线数据-白中班」：其他客服的会话，按上线时间倒序（白/中班以班次列区分）
- 「按客服汇总」：每天每个客服的分配工单数（去重）与记录数

样式：全表 Arial 10、时间列 yyyy-mm-dd hh:mm:ss、冻结首行；
明细页不按人分组（同一人的记录按时间与其他人自然穿插）。

班次判定（按「人 + 日期」逐天判定，班次随排班变动、不参考其他天的历史）：
- 每人每天的证据只有当天的登录 + 分配行为：上午（12:00 前）有活动 →
  当天白班（证据优先）；无上午活动但晚间（18:00 后）有活动 → 当天中班
  （中班 14:00-18:00、20:00-23:00，必然参与晚间时段）；只有下午活动时，
  踩中班上班点（14:00 前后 10 分钟内）开始的判中班（当天未结束、晚间
  证据尚未产生），更晚才出现的判白班（临时支援）；夜班名单固定夜班——
  非名单的人永远不判夜班。报表每次从全量数据重新生成，晚间数据到来后
  当天判定自动修正
- 分配数据 / 按客服汇总 / 上下线两页统一使用该「人+日期」班次
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DT_FORMAT = "yyyy\\-mm\\-dd\\ hh:mm:ss"
BASE_FONT = Font(name="Arial", size=10)

# 汇总页同一日期内的班次展示顺序：中班 → 白班 → 夜班
SUMMARY_SHIFT_ORDER = {"中班": 0, "白班": 1, "夜班": 2}

SHEET1_NAME = "分配数据"
SHEET1_HEADERS = ["时间", "工单ID", "班次", "客服", "队列类型", "ID"]
SHEET1_WIDTHS = {"A": 19.14, "B": 8.57, "C": 8.0, "D": 10.0, "E": 22.0, "F": 17.5}

SHEET2N_NAME = "agent上下线数据-夜班"
SHEET2D_NAME = "agent上下线数据-白中班"
SHEET2_HEADERS = ["班次", "客服姓名", "上线时间 (UTC+8)", "下线时间 (UTC+8)", "当前状态"]
SHEET2_WIDTHS = {"A": 8.0, "B": 16.96, "C": 19.14, "D": 19.14, "E": 16.5}

SHEET3_NAME = "按客服汇总"
SHEET3_HEADERS = ["日期", "班次", "客服", "分配工单数", "分配记录数"]
SHEET3_WIDTHS = {"A": 13.0, "B": 8.0, "C": 12.0, "D": 13.0, "E": 13.0}


def _shift_tz(dt_utc: datetime, tz_offset_hours: int) -> datetime:
    """UTC → 报表时区（默认 UTC+8）。"""
    return dt_utc + timedelta(hours=tz_offset_hours)


def _to_minutes(hhmm: str) -> int:
    """'08:30' → 510。"""
    hh, mm = hhmm.split(":")
    return int(hh) * 60 + int(mm)


class ShiftRules:
    """从 config.report 解析班次时段，提供按「一天内分钟数」投票的判定。

    时段模型：白班 day=[开始,结束]；中班 mid=多段 [开始,结束]（当前仅 19:00-23:00）。
    班次时间本身已含提前接单的容错，不再额外外扩。
    """

    def __init__(self, report_cfg: dict):
        self.night_names = {n.strip().lower() for n in report_cfg.get("night_shift_agents", [])}
        shifts = report_cfg.get("shifts", {})
        day = shifts.get("day", ["08:30", "18:00"])
        self.day_s = _to_minutes(day[0])
        self.day_e = _to_minutes(day[1])
        self.mid_periods = [(_to_minutes(s), _to_minutes(e))
                            for s, e in shifts.get("mid", [["19:00", "23:00"]])]
        self.mid_last_e = max((e for _, e in self.mid_periods), default=0)

    def is_fixed_night(self, name: str) -> bool:
        return name.strip().lower() in self.night_names

    def vote(self, m: int) -> str:
        """非名单客服按一天内分钟数判定：<= day_e 白班，其余中班。
        （夜班专属名单内的人，任何时间都只可能是白班/中班，不会标夜班。）"""
        if m <= self.day_e:
            return "白班"
        return "中班"

    def row_shift(self, name: str, m: int) -> str:
        """单条分配记录的班次：夜班名单固定为夜班，其余按时间落段（仅白/中）。"""
        if self.is_fixed_night(name):
            return "夜班"
        return self.vote(m)

    def per_day_shifts(self, assignment_rows: list[dict[str, Any]],
                       session_rows: list[dict[str, Any]],
                       tz_offset_hours: int,
                       schedule: dict[str, dict[str, str]] | None = None
                       ) -> dict[tuple[str, str], str]:
        """按「人 + 日期」逐天判定班次（班次随排班变动，不跨天累计历史）。

        判定优先级：
        1. 夜班名单 → 永远夜班（夜班固定，不依赖班表）；
        2. 班表命中（该人该日有「中」/「班」标记）→ 按班表；
           班表标「夜」的非名单人员忽略，继续走行为逻辑；
        3. 行为逻辑（当天登录 + 分配）：上午（12:00 前）有活动 → 白班
           （证据优先）；无上午但晚间（18:00 后）有活动 → 中班；仅下午
           活动时，踩中班上班点（中班开始后 10 分钟内，如 13:49）判中班，
           更晚出现判白班（临时支援）。报表每次从全量数据重新生成，
           晚间数据到来后当天判定自动修正。
        班表按日期粒度生效：没有该日期的月份自动整体回退行为逻辑。
        """
        NOON = 12 * 60
        mid1_s = self.mid_periods[0][0] if self.mid_periods else 14 * 60
        schedule = schedule or {}
        morning: set[tuple[str, str]] = set()
        evening: set[tuple[str, str]] = set()
        first_m: dict[tuple[str, str], int] = {}
        seen: set[tuple[str, str]] = set()
        labels: dict[tuple[str, str], str] = {}

        def collect(name: str, local) -> None:
            key = (name, local.strftime("%Y-%m-%d"))
            seen.add(key)
            if self.is_fixed_night(name):
                return
            m = local.hour * 60 + local.minute
            if m < NOON:
                morning.add(key)
            elif m > self.day_e:
                evening.add(key)
            if key not in first_m or m < first_m[key]:
                first_m[key] = m

        for r in session_rows:
            collect(r["agent_name"], _shift_tz(r["start_utc"], tz_offset_hours))
        for r in assignment_rows:
            collect(r["agent_name"], _shift_tz(r["event_date_utc"], tz_offset_hours))

        from schedule import lookup
        for key in seen:
            name, date_str = key
            if self.is_fixed_night(name):
                labels[key] = "夜班"
                continue
            scheduled = lookup(schedule, name, date_str)
            if scheduled in ("白班", "中班"):
                labels[key] = scheduled
            elif key in morning:
                labels[key] = "白班"
            elif key in evening:
                labels[key] = "中班"
            elif first_m.get(key, NOON) <= mid1_s + 10:
                labels[key] = "中班"  # 踩中班上班点开始 → 当天是中班
            else:
                labels[key] = "白班"  # 下午中途才出现 → 白班临时支援
        return labels


def _style_sheet(ws, widths: dict[str, float], autofilter: bool = False) -> None:
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.font = BASE_FONT
    ws.freeze_panes = "A2"
    if autofilter:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _write_session_sheet(ws, rows: list[dict[str, Any]],
                         day_shifts: dict[tuple[str, str], str],
                         tz_offset_hours: int) -> None:
    """写一个上下线工作表：班次|客服|上线|下线|状态。

    排序：上线日期倒序（最新日期最上）→ 同日期内客服姓名按字母表 →
    同人内上线时间倒序。班次为该客服在登录当天的班次（per_day_shifts）。
    """

    def sort_key(r: dict[str, Any]):
        local = _shift_tz(r["start_utc"], tz_offset_hours)
        return (-int(local.strftime("%Y%m%d")), r["agent_name"].lower(),
                -r["start_utc"].timestamp())

    ws.append(SHEET2_HEADERS)
    for r in sorted(rows, key=sort_key):
        ongoing = r["end_utc"] is None
        start_local = _shift_tz(r["start_utc"], tz_offset_hours)
        ws.append([
            day_shifts.get((r["agent_name"], start_local.strftime("%Y-%m-%d")), "白班"),
            r["agent_name"],
            start_local,
            "进行中..." if ongoing else _shift_tz(r["end_utc"], tz_offset_hours),
            "🟢 在线中" if ongoing else "已下线",
        ])
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=3).value is not None:
            ws.cell(row=row, column=3).number_format = DT_FORMAT
            if not isinstance(ws.cell(row=row, column=4).value, str):
                ws.cell(row=row, column=4).number_format = DT_FORMAT
    _style_sheet(ws, SHEET2_WIDTHS)


def build_report_xlsx(
    assignment_rows: list[dict[str, Any]],
    session_rows: list[dict[str, Any]],
    output_path: str,
    tz_offset_hours: int = 8,
    report_cfg: dict | None = None,
    now_utc: datetime | None = None,
) -> str:
    """生成报表 XLSX，返回路径。

    assignment_rows: store.query_assignment_since() 的返回值（时间为 UTC）
    session_rows:     store.query_sessions_since() 的返回值（时间为 UTC）
    report_cfg:       config.report 子字典（班次配置）
    now_utc:          预留（当前版本未使用）
    """
    report_cfg = report_cfg or {}
    rules = ShiftRules(report_cfg)

    # 班表兜底：存在则按 (人, 日期) 查班次；文件缺失/无该日期自动回退行为逻辑
    import schedule as schedule_mod
    schedule_map: dict[str, dict[str, str]] = {}
    schedule_path = report_cfg.get("schedule_file") or "客户专家班表.xlsx"
    if not os.path.isabs(schedule_path):
        schedule_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            schedule_path)
    if os.path.exists(schedule_path):
        try:
            schedule_map = schedule_mod.load_schedule(schedule_path)
        except Exception as e:  # noqa: BLE001
            logging.getLogger("rr.build").warning("班表解析失败，回退行为逻辑: %s", e)
    else:
        logging.getLogger("rr.build").info("未找到班表文件 %s，按行为逻辑判定", schedule_path)

    # 全表统一：班次按「人 + 日期」逐天判定（夜班名单 > 班表 > 行为逻辑），
    # 分配明细 / 按客服汇总 / 上下线两页共用
    day_shifts = rules.per_day_shifts(assignment_rows, session_rows,
                                      tz_offset_hours, schedule_map)

    def row_shift(name: str, local) -> str:
        """该客服在 local 当天的班次；无记录时退回按时间落段。"""
        label = day_shifts.get((name, local.strftime("%Y-%m-%d")))
        if label:
            return label
        return rules.row_shift(name, local.hour * 60 + local.minute)

    wb = Workbook()

    # ---- Sheet1 分配数据（整体时间倒序；班次为该客服当天的班次） ----
    ws1 = wb.active
    ws1.title = SHEET1_NAME
    ws1.append(SHEET1_HEADERS)
    for r in sorted(assignment_rows,
                    key=lambda r: (-r["event_date_utc"].timestamp(), -r["id"])):
        local = _shift_tz(r["event_date_utc"], tz_offset_hours)
        ws1.append([
            local,
            int(r["ticket_id"]) if str(r["ticket_id"]).isdigit() else r["ticket_id"],
            row_shift(r["agent_name"], local),
            r["agent_name"],
            r["queue_name"],
            r["id"],
        ])
    for row in range(2, ws1.max_row + 1):
        ws1.cell(row=row, column=1).number_format = DT_FORMAT
    _style_sheet(ws1, SHEET1_WIDTHS, autofilter=True)

    # ---- Sheet2 上下线数据拆两页：夜班 / 白中班（均按上线时间倒序） ----
    night_rows = [r for r in session_rows if rules.is_fixed_night(r["agent_name"])]
    other_rows = [r for r in session_rows if not rules.is_fixed_night(r["agent_name"])]
    _write_session_sheet(wb.create_sheet(SHEET2N_NAME), night_rows, day_shifts,
                         tz_offset_hours)
    _write_session_sheet(wb.create_sheet(SHEET2D_NAME), other_rows, day_shifts,
                         tz_offset_hours)

    # ---- Sheet3 按客服汇总（每天每客服每班次：去重工单数 + 记录数） ----
    ws3 = wb.create_sheet(SHEET3_NAME)
    summary: dict[tuple, dict[str, set]] = {}
    for r in assignment_rows:
        local = _shift_tz(r["event_date_utc"], tz_offset_hours)
        shift_label = row_shift(r["agent_name"], local)
        key = (local.strftime("%Y-%m-%d"), SUMMARY_SHIFT_ORDER.get(shift_label, 0),
               shift_label, r["agent_name"])
        agg = summary.setdefault(key, {"tickets": set(), "count": 0})
        agg["tickets"].add(str(r["ticket_id"]))
        agg["count"] += 1
    ws3.append(SHEET3_HEADERS)
    # 日期倒序（最新在上）→ 班次 中→白→夜 → 工单量降序 → 客服名
    for key in sorted(summary, key=lambda k: (-int(k[0].replace("-", "")), k[1],
                                              -len(summary[k]["tickets"]), k[3])):
        date_str, _, shift_label, name = key
        agg = summary[key]
        ws3.append([date_str, shift_label, name, len(agg["tickets"]), agg["count"]])
    _style_sheet(ws3, SHEET3_WIDTHS, autofilter=True)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> int:
    """命令行自测：从库里取数生成表格到 output/。"""
    import argparse
    import sys
    from datetime import timezone

    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import config_utils  # noqa: E402
    import store  # noqa: E402

    parser = argparse.ArgumentParser(description="生成 RR 报表 XLSX")
    parser.add_argument("--output", "-o", required=True, help="输出 .xlsx 路径")
    args = parser.parse_args()

    config = config_utils.load_config()
    now_utc = datetime.now(timezone.utc).replace(tzinfo=None)
    a_rows = store.query_assignment_since(
        int(config["report"]["assignment_window_hours"]), now_utc)
    s_rows = store.query_sessions_since(
        int(config["report"]["availability_days"]), now_utc)
    build_report_xlsx(a_rows, s_rows, args.output,
                      tz_offset_hours=int(config["rr"].get("tz_offset_hours", 8)),
                      report_cfg=config["report"], now_utc=now_utc)
    print(f"已生成: {args.output}（分配 {len(a_rows)} 条 / 会话 {len(s_rows)} 条）")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
