"""读取《客户专家班表.xlsx》，提供 (客服, 日期) → 班次 的查询。

班表格式（每月一个工作表，任意表名均可，按内容识别）：
- 前 6 行里存在一行「日期行」：该行有 >=3 个 datetime 单元格
- A 列自日期行的下一行起为人名
- 单元格值：「中」→ 中班、「班」→ 白班、「夜」→ 夜班；
  休 / 假 / 1/2假 / 下午假 等不含班次信息，跳过（该人该日回退行为逻辑）

名字匹配忽略大小写，并支持前缀互配（如班表「Vv」可匹配数据「vv.deng」）。
按日期粒度查询：班表里没有该日期（或该人该日无有效班次标记）即视为未覆盖。
"""

from __future__ import annotations

import logging
from datetime import datetime, timedelta
from typing import Any

log = logging.getLogger("rr.schedule")

SCHEDULE_VALUE_MAP = {"中": "中班", "班": "白班", "夜": "夜班"}


def load_schedule(path: str) -> dict[str, dict[str, str]]:
    """解析班表，返回 {日期str: {人名(小写): 班次}}。

    只读取「YYYY年M月」命名的正式月表（带括号后缀的作废/副本/预安排表忽略）。
    日期以表标题的年月为准：单元格日期落在该月 ±7 天内直接采用；
    若恰好早一年（模板复制后日期未改），自动按 +1 年对齐采用；
    其余（更早/更晚）视为无效，整表无有效日期则跳过。
    """
    import re

    import openpyxl

    result: dict[str, dict[str, str]] = {}

    def aligned(d: datetime, lo: datetime, hi: datetime) -> datetime | None:
        if lo <= d <= hi:
            return d
        try:
            d2 = d.replace(year=d.year + 1)
        except ValueError:  # 2/29
            return None
        return d2 if lo <= d2 <= hi else None

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        m = re.match(r"^(\d{4})年(\d{1,2})月$", str(ws.title).strip())
        if not m:
            continue
        year, month = int(m.group(1)), int(m.group(2))
        first = datetime(year, month, 1)
        nxt = datetime(year + (month == 12), month % 12 + 1, 1)
        lo = first - timedelta(days=7)
        hi = nxt - timedelta(days=1) + timedelta(days=7)
        date_cols: dict[int, str] = {}
        name_start_row = 0
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True),
                                     start=1):
            dts = {}
            for i, c in enumerate(row, start=1):
                if isinstance(c, datetime):
                    d = aligned(c, lo, hi)
                    if d is not None:
                        dts[i] = d
            if len(dts) >= 3:
                date_cols = {i: c.strftime("%Y-%m-%d") for i, c in dts.items()}
                name_start_row = r_idx + 1
                break
        if not date_cols:
            log.info("班表工作表 [%s]: 无有效日期，跳过", ws.title)
            continue  # 非排班工作表或日期无效
        n_rows = 0
        for row in ws.iter_rows(min_row=name_start_row, values_only=True):
            if not row:
                continue
            name = str(row[0]).strip() if row[0] is not None else ""
            if not name:
                continue
            name_l = name.lower()
            for col_idx, date_str in date_cols.items():
                v = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if v is None:
                    continue
                shift = SCHEDULE_VALUE_MAP.get(str(v).strip())
                if shift:
                    result.setdefault(date_str, {})[name_l] = shift
            n_rows += 1
        log.info("班表工作表 [%s]: %d 个日期列, %d 行人员", ws.title,
                 len(date_cols), n_rows)
    wb.close()
    total = sum(len(v) for v in result.values())
    log.info("班表解析完成: %d 天 / %d 人日条目", len(result), total)
    return result


def lookup(schedule: dict[str, dict[str, str]], name: str,
           date_str: str) -> str | None:
    """查询某客服某日的班次；无匹配返回 None。

    先精确匹配（忽略大小写），再前缀互配（最短 2 字符）。
    """
    day = schedule.get(date_str)
    if not day:
        return None
    n = name.strip().lower()
    if n in day:
        return day[n]
    best: str | None = None
    best_len = 0
    for sname, shift in day.items():
        if (sname.startswith(n) or n.startswith(sname)) and min(len(sname), len(n)) >= 2:
            if len(sname) > best_len:  # 取最长的前缀匹配，更精确
                best, best_len = shift, len(sname)
    return best
